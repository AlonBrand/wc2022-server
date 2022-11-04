from openpyxl import load_workbook
import pandas as pd

def insert_row(file_name, row_data):
    try:
        wb = load_workbook("files/{file_name}.xlsx".format(file_name=file_name))
        ws = wb.worksheets[0]
        ws.append(row_data)
        wb.save("files/{file_name}.xlsx".format(file_name=file_name))
        return True
    except:
        return False

def get_table(file_name):
    games = []
    wb = load_workbook("files/{file_name}.xlsx".format(file_name=file_name))
    ws = wb.worksheets[0]
    all_rows = list(ws.rows)

    for row in all_rows[1:]:
        curr_game = []
        for cell in row:
            curr_game.append(cell.value)
        games.append(curr_game)

    return games

def search_in_table(file_name, user_name, user_password):
    wb = load_workbook("files/{file_name}.xlsx".format(file_name=file_name))
    ws = wb.worksheets[0]
    all_rows = list(ws.rows)

    for row in all_rows[1:]:
        if row[0].value == user_name:
            if row[1].value == user_password:
                return 'User connected'
            elif row[1].value != user_password:
                return 'Wrong password'
        else: continue

    return "User not exist!"
    